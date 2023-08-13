from datetime import datetime
from websocket import create_connection
import requests
import json
import random
import re
import string
import openpyxl
import pytz
import os

# Set the timezone to Indian Standard Time
IST = pytz.timezone('Asia/Kolkata')
today = datetime.now(IST)
d1 = today.strftime("%d-%m-%Y")


#Function for creating excel file with today's date as name
def createExcelFile():
  global IST, d1
  # Update with your desired filename or use a variable to dynamically set the filename
  filepath = f"Output/{d1}.xlsx"
  log_file = "restartLog.txt"
  if not os.path.exists(filepath):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if ws.max_row == 0:
      ws.append(["Datetime", "Price"])
    wb.save(filepath)
    print(f"Excel file '{d1}.xlsx' created successfully.")
  else:
    print(f"Excel file '{d1}.xlsx' already exists. Skipping creation.")
    with open(log_file, "a") as log:
      log.write(f"{datetime.now(IST)}: Restarted '{d1}.xlsx'\n")


def generateSession():
  stringLength = 12
  letters = string.ascii_lowercase
  random_string = "".join(random.choice(letters) for i in range(stringLength))
  return "qs_" + random_string


def prependHeader(st):
  return "~m~" + str(len(st)) + "~m~" + st


def constructMessage(func, paramList):
  return json.dumps({"m": func, "p": paramList}, separators=(",", ":"))


def createMessage(func, paramList):
  return prependHeader(constructMessage(func, paramList))


def sendMessage(ws, func, args):
  ws.send(createMessage(func, args))


def sendPingPacket(ws, result):
  pingStr = re.findall(".......(.*)", result)
  if len(pingStr) != 0:
    pingStr = pingStr[0]
    ws.send("~m~" + str(len(pingStr)) + "~m~" + pingStr)


def socketJob(ws):
  while True:
    try:
      result = ws.recv()
      if "quote_completed" in result or "session_id" in result:
        continue
      Res = re.findall("^.*?({.*)$", result)
      if len(Res) != 0:
        jsonRes = json.loads(Res[0])

        if jsonRes["m"] == "qsd":
          symbol = jsonRes["p"][1]["n"]
          price = jsonRes["p"][1]["v"]["lp"]
          timestamp = jsonRes["p"][1]["v"]["lp_time"]
          time = datetime.fromtimestamp(timestamp)
          print(f"{symbol} -> {price} : {time}")
      else:
        # ping packet
        sendPingPacket(ws, result)
    except KeyboardInterrupt:
      print("\nGoodbye!")
      exit(0)
    except Exception as e:
      #print(f"ERROR: {e}\nTradingView message: {result}")
      continue


def socketJobNew(ws):
  workbook = openpyxl.load_workbook(f'Output/{d1}.xlsx')
  worksheet = workbook['Sheet1']
  while True:
    try:
      result = ws.recv()
      if "quote_completed" in result or "session_id" in result:
        continue
      Res = re.findall("^.*?({.*)$", result)
      if len(Res) != 0:
        Res1 = re.split(r'~m~\d+~m~', result)
        length = len(Res1) - 1
        for i in range(1, length + 1):
          # if (length > 1):
          #   print(f"Extra data but managed: {Res1}")
          jsonRes = json.loads(Res1[i])
          if jsonRes["m"] == "qsd":
            symbol = jsonRes["p"][1]["n"]
            price = jsonRes["p"][1]["v"]["lp"]
            timestamp = jsonRes["p"][1]["v"]["lp_time"]
            time = datetime.fromtimestamp(timestamp)
            data = [time, price]
            worksheet.append(data)
            workbook.save(f'Output/{d1}.xlsx')
            print(f"{symbol} -> {price} : {time}")
      else:
        # ping packet
        sendPingPacket(ws, result)
    except KeyboardInterrupt:
      print("\nGoodbye!")
      exit(0)
    except Exception as e:
      try:
        if 'lp_time' in str(e):
          # print('no lp time')
          if len(Res) != 0:
            Res1 = re.split(r'~m~\d+~m~', result)
            length = len(Res1) - 1
            for i in range(1, length + 1):
              # if (length > 1):
              #   print(f"Extra data but managed: {Res1}")
              jsonRes = json.loads(Res1[i])
              if jsonRes["m"] == "qsd":
                symbol = jsonRes["p"][1]["n"]
                price = jsonRes["p"][1]["v"]["lp"]
                time = datetime.now(IST).strftime(r'%Y-%m-%d %H:%M:%S')
                data = [time, price]
                worksheet.append(data)
                workbook.save(f'Output/{d1}.xlsx')
                print(f"{symbol} -> {price} : {time}")
      except Exception as f:
        # print(f"ERROR: {f}\nTradingView message: {result}")
        continue
      else:
        # print(f"ERROR: {e}\nTradingView message: {result}")
        continue


def getSymbolId():
  # data = search(pair, market)
  # symbol_name = data["symbol"]
  # try:
  #     broker = data["prefix"]
  # except KeyError:
  #     broker = data["exchange"]
  # symbol_id = f"{broker.upper()}:{symbol_name.upper()}"
  # print(symbol_id, end="\n\n")
  symbol_id = 'NSE:NIFTY'
  print(symbol_id)
  return symbol_id


def main():

  # serach btcusdt from crypto category
  symbol_id = getSymbolId()

  # create tunnel
  tradingViewSocket = "wss://data.tradingview.com/socket.io/websocket"
  headers = json.dumps({"Origin": "https://data.tradingview.com"})
  ws = create_connection(tradingViewSocket, headers=headers)
  session = generateSession()

  # Send messages
  sendMessage(ws, "quote_create_session", [session])
  sendMessage(ws, "quote_set_fields", [session, "lp", "lp_time"])
  sendMessage(ws, "quote_add_symbols", [session, symbol_id])

  # Start job
  socketJobNew(ws)

  def websocketClose():
    ws.close()
